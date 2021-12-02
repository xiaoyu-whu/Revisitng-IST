import time
import os
import pandas as pd
import openpyxl

if __name__ == '__main__':

    folder_path = '/Users/xiao/Documents/project/Python/Revisiting-IST/test_result/'
    print("start time：", time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
    starttime = time.time()

    new_files = []
    for root, dirs, files, in os.walk(folder_path):
        for file in files:
            if file.endswith('.xlsx'):
                new_files.append(file)
        new_files.sort()
        print(new_files)

        filenumber=int(len(new_files) / 2)
        print(filenumber)
        for i in range(0, filenumber):
            file_R = os.path.join(folder_path, new_files[i])
            file_sk = os.path.join(folder_path, new_files[i+filenumber])
            last_name = new_files[i][8:]

            print('file_R:',file_R)
            print('file_sk',file_sk)
            print(last_name)

            wb_R = openpyxl.load_workbook(file_R)  # 读取xlsx
            wb_sk = openpyxl.load_workbook(file_sk)
            sheetnames_train = wb_R.get_sheet_names()  # 获取所有表
            sheetnames_test = wb_sk.get_sheet_names()
            ws_train = wb_R.get_sheet_by_name(sheetnames_train[0])  # 获取第一个表
            ws_test = wb_sk.get_sheet_by_name(sheetnames_test[0])


            dataset_R = pd.DataFrame(ws_train.values)
            dataset_sk = pd.DataFrame(ws_test.values)

            merge_pd = pd.concat([dataset_R,dataset_sk],axis=1,ignore_index=True)
            merge_pd = merge_pd.drop(6,axis=1)

            if not os.path.exists("/Users/xiao/Documents/project/Python/Revisiting-IST/merged_result"):
                os.makedirs("/Users/xiao/Documents/project/Python/Revisiting-IST/merged_result")
            merge_pd.to_excel('/Users/xiao/Documents/project/Python/Revisiting-IST/merged_result/merged'+last_name,index=False,header=None)
            print(i)


    print("end time：", time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
    endtime = time.time()
    print('cost time:', endtime - starttime, 's')